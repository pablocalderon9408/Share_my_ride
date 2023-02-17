
from openpyxl import load_workbook
import os

# Django
from django.conf import settings
from django.core.management.base import BaseCommand


class Command(BaseCommand):
    def handle(self, *args, **options):
        file = os.path.join(settings.APPS_DIR, 'circles/management/commands/Activar Usuario en Ruta.xlsx')
        workbook = load_workbook(filename=file)
        worksheet = workbook['Sheet1']
        all_data = []

        for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=55):
            for cell in row:
                all_data.append(cell.value)
        print(all_data)
        print(len(all_data))